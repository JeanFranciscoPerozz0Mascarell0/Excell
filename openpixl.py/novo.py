from tkinter import *
from tkinter import messagebox
import openpyxl

def limpar_formulario():
    campo_nome.delete(0, END)
    entrada_idade.delete(0, END)

def criar_planilha():
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    nome = campo_nome.get()
    idade = int(entrada_idade.get())
    if idade < 0 or idade > 120:
        messagebox.showerror("ERRO", "A idade deve estar entre 0 a 120.")
        return
    ws.append([nome, idade])
    wb.save('planilha1.xlsx')

def ler_planilha():
    wb = openpyxl.load_workbook('planilha1.xlsx')
    ws = wb.worksheets[0]
    nome = ws.cell(row=1, column='1').value
    idade = ws.cell(row=1, column='2').value

    campo_nome.delete(0, END)
    campo_nome.insert(0, nome)

    entrada_idade.delete(0, END)
    entrada_idade.insert(0, END)

def ler_planilha_mostrar():
    wb = openpyxl.load_workbook("planilha.xlsx")
    ws = wb.worksheets[0]
    messagebox.showinfo("Dados da Planilha", f"Nome: {nome}\nIdade: {idade}")

janela = Tk()
janela.title("Tela de Cadastro")
janela.geometry("900x300")
Label(janela, text="Manipulando valores de duas células", font="Arial 16 bold").pack
nome = Label(janela, text="Nome:", font="Georgia").place(x=30, y=50)
campo_nome = Entry(janela, width=50, font="Verdana")
campo_nome.place(x=150, y=50)

idade = Label(janela, text="Idade:", font="Georgia").place(x=30, y=90)
campo_idade = Entry(janela, width=50, font="Verdana")
campo_idade.place(x=150, y=95)

botao = Button(janela, text="Limpar", command=limpar_formulario).place(x=100, y=170)
botao2 = Button(janela, text="Criar planilha", command=criar_planilha).place(x=10, y=170)
botao3 = Button(janela, text="Ler Planilha e inserir nos campos", command=ler_planilha).place(x=160, y=170)
botao4 = Button(janela, text="Ler planilha e mostrar na tela", command=ler_planilha_mostrar).place(x=355, y=170)

janela.mainloop()