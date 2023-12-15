import openpyxl

df1_data = [2, 4, 6, 8]
df2_data = [100, 150, 200, 250]
df3_data = [3, 6, 9, 12]

wb = openpyxl.Workbook()
ws = wb.worksheets[0]

ws1 = wb.create_sheet()
ws2 = wb.create_sheet()
ws3 = wb.create_sheet()
ws.append([2, 4, 6, 8])
ws1.append(df1_data)
ws2.append(df2_data)
ws3.append(df3_data)

ws1.title = 'Tabela 1'
ws2.title = 'Tabela 2'
ws3.title = 'Tabela 3'


wb.save('tabelas_exemplo.xlsx')