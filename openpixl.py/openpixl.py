import openpyxl
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
import datetime as dt

hoje = dt.datetime.now()
formatado = "%Y-%m-%d %h:%M"
hojeeditado = hoje.strftime(formatado).replace(":", "")
wb = openpyxl.Workbook()
ws = wb.worksheets[0]
ws.append(["Nome", "Idade"])
ws.append(["Lucas", 30])

for i in range(1, 11):
    ws['A' + str(i)] = i

a1 = ws['A1']
b1 = ws['B1']
ft = Font(color = "FF0000", italic=True)
b1.font = Font(color="FF0000", bold=True)
a1.font = ft

wb.save("teste.xlsx")