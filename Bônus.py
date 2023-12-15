from tkinter import *
from tkinter import messagebox
import openpyxl

def limpar_formulario():
    campo_nome.delete(0,END)
    entrada_idade.delete(0, END)

def criar_planilha():
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    nome = campo_nome.get()
    idade = int(entrada_idade.get())
    if idade < 0 or idade > 120:
        messagebox.showerror("ERRO", "A idade deve estar entre 0 a 120.")
        return
    ws.append([nome, idade])
    wb.save('planilha1.xlsx')

def ler_planilha():
    wb = openpyxl.load_workbook('planilha1.xlsx')
    ws = wb.worksheets[0]
    nome = ws.cell(row=1,column=1).value
    idade = ws.cell(row=1,column=2).value

    campo_nome.delete(0,END)
    campo_nome.insert(0,nome)
    entrada_idade.delete(0,END)
    entrada_idade.insert(0,idade)

def ler_planilha_mostrar():
    wb = openpyxl.load_workbook("planilha1.xlsx")
    ws = wb.worksheets[0]
    nome = ws.cell(row=1, column=1).value
    idade = ws.cell(row=1, column=2).value
    messagebox.showinfo("Dados da Planilha", f"Nome: {nome}\nIdade: {idade}")
Janela = Tk()
Janela.title("Tela de Cadastro")
Janela.geometry("900x300")
Label(Janela, text='Manipulando valores de duas células',font="Arial 16 bold").pack()
nome = Label(Janela, text="Nome", font="Georgia").place(x=30, y=50)
campo_nome = Entry(Janela,width=50, font="Verdana")
campo_nome.place(x=150, y=50)
idade = Label(Janela, text="Idade",font="Georgia").place(x=30, y=90)
entrada_idade = Entry(Janela, width=50, font="Verdana")
entrada_idade.place(x=150, y=90)

Button(Janela, text="Limpar", command=limpar_formulario).place(x=100, y=170)
Button(Janela, text="Criar Planilha", command=criar_planilha).place(x=10, y=170)
Button(Janela, text="Ler Planilha e inserir nos campos", command=ler_planilha).place(x=160, y=170)
Button(Janela, text="Ler planilha e mostrar na tela", command=ler_planilha_mostrar).place(x=355, y=170)

Janela.mainloop()
